import asyncio
import random
import re
import os
import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
from fake_useragent import UserAgent
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import httpx


# ─────────────────────────────────────────────────────────────
# TEXT CLEANER
# ─────────────────────────────────────────────────────────────
def clean_text(value: str) -> str:
    if not value or str(value).strip() in ("", "N/A", "None"):
        return "N/A"
    value = str(value)
    value = value.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    value = re.sub(r"\s+", " ", value).strip()
    return value if value else "N/A"


# ─────────────────────────────────────────────────────────────
# LEAD SCRAPER
# ─────────────────────────────────────────────────────────────
class LeadScraper:
    def __init__(self, headless: bool = False):
        """
        headless=False  → browser window opens so you can SEE what's happening
        headless=True   → runs invisibly in background (use after it's working)
        """
        self.ua     = UserAgent()
        self.leads  = []
        self.headless = headless

    async def init_browser(self, playwright):
        browser = await playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--start-maximized",
            ]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 800},
            extra_http_headers={
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            }
        )
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        """)
        return browser, context

    async def human_delay(self, min_sec=1.5, max_sec=3.5):
        await asyncio.sleep(random.uniform(min_sec, max_sec))

    async def slow_scroll(self, page, times=5):
        for _ in range(times):
            await page.evaluate("window.scrollBy(0, 300)")
            await asyncio.sleep(random.uniform(0.4, 0.9))

    # ──────────────────────────────────────────────────────────
    # SCRAPER 1: GOOGLE MAPS (most reliable for Indian cities)
    # ──────────────────────────────────────────────────────────
    async def scrape_google_maps(
        self,
        query: str,
        location: str,
        max_results: int = 20
    ):
        print(f"\n{'='*55}")
        print(f"  [Google Maps] '{query}' in '{location}'")
        print(f"{'='*55}")

        async with async_playwright() as p:
            browser, context = await self.init_browser(p)
            page = await context.new_page()

            search_term = f"{query} in {location}"
            url = f"https://www.google.com/maps/search/{search_term.replace(' ', '+')}"

            print(f"  → Opening: {url}")
            await page.goto(url, wait_until="domcontentloaded", timeout=45000)
            await page.wait_for_timeout(3000)  # give Maps 3 extra seconds to render
            await self.human_delay(3, 5)

            # ── Accept cookies popup if it appears (EU/some regions) ──
            try:
                accept_btn = await page.query_selector("button[aria-label='Accept all']")
                if accept_btn:
                    await accept_btn.click()
                    await self.human_delay(1, 2)
                    print("  → Accepted cookies popup")
            except Exception:
                pass

            # ── Wait for results feed to appear ───────────────────────
            print("  → Waiting for results to load...")
            try:
                await page.wait_for_selector(
                    "div[role='feed']",
                    timeout=15000
                )
                print("  → Results feed found ✓")
            except Exception:
                print("  ✗ Results feed not found — trying alternative...")
                # Take screenshot so you can see what happened
                await page.screenshot(path=os.path.expanduser("~/Desktop/debug_maps.png"))
                print("  → Screenshot saved to Desktop/debug_maps.png")
                await browser.close()
                return

            scraped = 0
            scroll_attempts = 0
            MAX_SCROLL_ATTEMPTS = 15

            while scraped < max_results and scroll_attempts < MAX_SCROLL_ATTEMPTS:

                # ── Get all listing links currently visible ────────────
                # Google Maps uses multiple possible selectors — try all
                listings = await page.query_selector_all("a.hfpxzc")

                if not listings:
                    listings = await page.query_selector_all(
                        "div[role='feed'] a[href*='/maps/place/']"
                    )

                print(f"  → Found {len(listings)} listings on screen, scraped {scraped} so far")

                for listing in listings:
                    if scraped >= max_results:
                        break

                    try:
                        # Click the listing to open its detail panel
                        await listing.click()
                        await self.human_delay(2, 3)

                        # ── Try multiple selector patterns for each field ──

                        # NAME — try 3 different selectors
                        name = "N/A"
                        for sel in ["h1.DUwDvf", "h1[class*='fontHeadlineLarge']", "h1"]:
                            el = await page.query_selector(sel)
                            if el:
                                name = await el.inner_text()
                                break

                        # PHONE — try multiple patterns
                        phone = "N/A"
                        for sel in [
                            "button[data-item-id^='phone:tel']",
                            "button[aria-label*='Phone']",
                            "[data-tooltip*='+']",
                            "span[aria-label*='phone' i]",
                        ]:
                            el = await page.query_selector(sel)
                            if el:
                                phone = (
                                    await el.get_attribute("aria-label")
                                    or await el.get_attribute("data-item-id")
                                    or await el.inner_text()
                                )
                                # Clean up phone — extract digits and +
                                phone = re.sub(r"[^\d\+\-\s\(\)]", "", phone).strip()
                                if phone:
                                    break

                        # ADDRESS
                        address = "N/A"
                        for sel in [
                            "button[data-item-id='address']",
                            "button[aria-label*='Address']",
                            "[data-item-id*='address']",
                        ]:
                            el = await page.query_selector(sel)
                            if el:
                                address = (
                                    await el.get_attribute("aria-label")
                                    or await el.inner_text()
                                )
                                address = address.replace("Address: ", "").strip()
                                break

                        # WEBSITE
                        website = "N/A"
                        for sel in [
                            "a[data-item-id='authority']",
                            "a[aria-label*='website' i]",
                            "a[href*='http'][aria-label]",
                        ]:
                            el = await page.query_selector(sel)
                            if el:
                                website = await el.get_attribute("href") or "N/A"
                                break

                        # RATING
                        rating = "N/A"
                        for sel in [
                            "div.F7nice span[aria-hidden='true']",
                            "span[aria-label*='stars' i]",
                            "div.fontBodyMedium span[aria-hidden]",
                        ]:
                            el = await page.query_selector(sel)
                            if el:
                                rating = await el.inner_text()
                                if re.search(r"\d", rating):
                                    break
                                rating = "N/A"

                        # CATEGORY
                        category = "N/A"
                        for sel in [
                            "button.DkEaL",
                            "span.YhemCb",
                            "button[jsaction*='category']",
                        ]:
                            el = await page.query_selector(sel)
                            if el:
                                category = await el.inner_text()
                                break

                        # ── Skip if we got nothing useful ─────────────
                        if name == "N/A":
                            continue

                        # ── Check for duplicate ───────────────────────
                        already_scraped = any(
                            l["business_name"] == clean_text(name)
                            for l in self.leads
                        )
                        if already_scraped:
                            continue

                        lead = {
                            "source":        "google_maps",
                            "business_name": clean_text(name),
                            "phone":         clean_text(phone),
                            "address":       clean_text(address),
                            "website":       clean_text(website),
                            "rating":        clean_text(rating),
                            "category":      clean_text(category) if category != "N/A" else query,
                            "email":         "N/A",
                        }

                        self.leads.append(lead)
                        scraped += 1
                        print(
                            f"     [{scraped}/{max_results}] "
                            f"{lead['business_name'][:35]:<35} | "
                            f"{lead['phone']:<15} | "
                            f"{lead['address'][:30]}"
                        )

                    except Exception as e:
                        print(f"     → Skipped listing: {e}")
                        continue

                # ── Scroll results panel to load more ────────────────
                try:
                    feed = await page.query_selector("div[role='feed']")
                    if feed:
                        await feed.evaluate("el => el.scrollTop += 1500")
                    scroll_attempts += 1
                    await self.human_delay(2, 3)
                except Exception:
                    scroll_attempts += 1

            print(f"\n  ✅ Google Maps done — {scraped} leads collected")
            await browser.close()

    # ──────────────────────────────────────────────────────────
    # SCRAPER 2: JUSTDIAL (best for Indian businesses)
    # ──────────────────────────────────────────────────────────
    async def scrape_justdial(
        self,
        query: str,
        location: str,
        max_results: int = 20
    ):
        print(f"\n{'='*55}")
        print(f"  [JustDial] '{query}' in '{location}'")
        print(f"{'='*55}")

        async with async_playwright() as p:
            browser, context = await self.init_browser(p)
            page = await context.new_page()

            # Format: https://www.justdial.com/Mumbai/Restaurants
            loc   = location.replace(" ", "-")
            q     = query.replace(" ", "-").title()
            url   = f"https://www.justdial.com/{loc}/{q}"

            print(f"  → Opening: {url}")

            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=45000)
                await self.human_delay(3, 5)

                # Close login popup if appears
                try:
                    close_btn = await page.query_selector(
                        "span.css-jgqnru, button.close, [aria-label='Close']"
                    )
                    if close_btn:
                        await close_btn.click()
                        await self.human_delay(1, 2)
                        print("  → Closed popup")
                except Exception:
                    pass

                scraped = 0
                scroll_rounds = 0

                while scraped < max_results and scroll_rounds < 10:

                    html  = await page.content()
                    soup  = BeautifulSoup(html, "html.parser")

                    # JustDial listing cards
                    cards = soup.select("li.cntanr, li[class*='store-cards'], div.resultbox_info")

                    print(f"  → Found {len(cards)} cards | scraped {scraped} so far")

                    for card in cards:
                        if scraped >= max_results:
                            break

                        # NAME
                        name_el = (
                            card.select_one("span.lng_cont_name")
                            or card.select_one("a.store-name")
                            or card.select_one("h2 a")
                            or card.select_one(".resultbox_title_anchor")
                        )

                        # PHONE — JustDial hides numbers, grab what's visible
                        phone_el = (
                            card.select_one("p.contact-info")
                            or card.select_one("span.contact_info")
                            or card.select_one("div.telCntct")
                            or card.select_one("[class*='contact']")
                        )

                        # ADDRESS
                        addr_el = (
                            card.select_one("span.cont_fl_addr")
                            or card.select_one("p.jd_address")
                            or card.select_one("[class*='address']")
                        )

                        # RATING
                        rating_el = (
                            card.select_one("span.green-box")
                            or card.select_one("[class*='rating']")
                        )

                        name   = name_el.text.strip()   if name_el   else "N/A"
                        phone  = phone_el.text.strip()  if phone_el  else "N/A"
                        addr   = addr_el.text.strip()   if addr_el   else "N/A"
                        rating = rating_el.text.strip() if rating_el else "N/A"

                        if name == "N/A":
                            continue

                        already = any(
                            l["business_name"] == clean_text(name)
                            for l in self.leads
                        )
                        if already:
                            continue

                        lead = {
                            "source":        "justdial",
                            "business_name": clean_text(name),
                            "phone":         clean_text(phone),
                            "address":       clean_text(addr),
                            "website":       "N/A",
                            "rating":        clean_text(rating),
                            "category":      query,
                            "email":         "N/A",
                        }
                        self.leads.append(lead)
                        scraped += 1
                        print(
                            f"     [{scraped}/{max_results}] "
                            f"{lead['business_name'][:35]:<35} | "
                            f"{lead['phone']}"
                        )

                    # Scroll down to load more
                    await page.evaluate("window.scrollBy(0, 1500)")
                    scroll_rounds += 1
                    await self.human_delay(2, 3)

                print(f"\n  ✅ JustDial done — {scraped} leads collected")

            except Exception as e:
                print(f"  ✗ JustDial error: {e}")
                await page.screenshot(
                    path=os.path.expanduser("~/Desktop/debug_justdial.png")
                )
                print("  → Screenshot saved to Desktop/debug_justdial.png")
            finally:
                await browser.close()

    # ──────────────────────────────────────────────────────────
    # SCRAPER 3: INDIAMART (B2B suppliers & manufacturers)
    # ──────────────────────────────────────────────────────────
    async def scrape_indiamart(
        self,
        query: str,
        location: str = "",
        max_results: int = 20
    ):
        print(f"\n{'='*55}")
        print(f"  [IndiaMart] '{query}'")
        print(f"{'='*55}")

        async with async_playwright() as p:
            browser, context = await self.init_browser(p)
            page = await context.new_page()

            search = f"{query} {location}".strip().replace(" ", "+")
            url = f"https://dir.indiamart.com/search.mp?ss={search}"

            print(f"  → Opening: {url}")
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=45000)
                await self.human_delay(3, 5)

                scraped = 0
                scroll_rounds = 0

                while scraped < max_results and scroll_rounds < 8:
                    html = await page.content()
                    soup = BeautifulSoup(html, "html.parser")

                    cards = soup.select(
                        "div.card, div.prd-card, div.company-listing, div[class*='supplier']"
                    )

                    print(f"  → Found {len(cards)} cards | scraped {scraped} so far")

                    for card in cards:
                        if scraped >= max_results:
                            break

                        name_el  = (
                            card.select_one("h2 a")
                            or card.select_one(".company-name a")
                            or card.select_one("a.companyname")
                        )
                        phone_el = (
                            card.select_one(".phone")
                            or card.select_one("[class*='phone']")
                            or card.select_one("[class*='contact']")
                        )
                        addr_el  = (
                            card.select_one(".company-address")
                            or card.select_one("[class*='address']")
                            or card.select_one("span.loc")
                        )

                        name  = name_el.text.strip()  if name_el  else "N/A"
                        phone = phone_el.text.strip() if phone_el else "N/A"
                        addr  = addr_el.text.strip()  if addr_el  else "N/A"

                        if name == "N/A":
                            continue

                        already = any(
                            l["business_name"] == clean_text(name)
                            for l in self.leads
                        )
                        if already:
                            continue

                        lead = {
                            "source":        "indiamart",
                            "business_name": clean_text(name),
                            "phone":         clean_text(phone),
                            "address":       clean_text(addr),
                            "website":       "N/A",
                            "rating":        "N/A",
                            "category":      query,
                            "email":         "N/A",
                        }
                        self.leads.append(lead)
                        scraped += 1
                        print(
                            f"     [{scraped}/{max_results}] "
                            f"{lead['business_name'][:35]:<35} | "
                            f"{lead['phone']}"
                        )

                    await page.evaluate("window.scrollBy(0, 1500)")
                    scroll_rounds += 1
                    await self.human_delay(2, 3)

                print(f"\n  ✅ IndiaMart done — {scraped} leads collected")

            except Exception as e:
                print(f"  ✗ IndiaMart error: {e}")
            finally:
                await browser.close()

    # ──────────────────────────────────────────────────────────
    # EMAIL FINDER via Hunter.io API
    # ──────────────────────────────────────────────────────────
    async def enrich_leads_with_emails(self, hunter_api_key: str):
        eligible = [
            l for l in self.leads
            if l.get("website", "N/A") != "N/A"
        ]
        print(f"\n[Hunter.io] Enriching {len(eligible)} leads that have websites...")

        for i, lead in enumerate(self.leads):
            if lead.get("website", "N/A") == "N/A":
                continue
            domain = (
                lead["website"]
                .replace("https://", "")
                .replace("http://", "")
                .split("/")[0]
            )
            try:
                async with httpx.AsyncClient() as client:
                    resp = await client.get(
                        "https://api.hunter.io/v2/domain-search",
                        params={"domain": domain, "api_key": hunter_api_key},
                        timeout=5.0
                    )
                    emails = resp.json().get("data", {}).get("emails", [])
                    if emails:
                        best = sorted(
                            emails,
                            key=lambda x: x.get("confidence", 0),
                            reverse=True
                        )
                        self.leads[i]["email"] = best[0].get("value", "N/A")
                        print(f"  ✓ {self.leads[i]['email']}  ← {lead['business_name']}")
            except Exception:
                pass
            await asyncio.sleep(0.5)

    # ──────────────────────────────────────────────────────────
    # EXPORT — formatted Excel + clean CSV
    # ──────────────────────────────────────────────────────────
    def export(self, filename_base: str = None):
        if not self.leads:
            print("\n  ✗ No leads collected. Check the debug screenshots on your Desktop.")
            return

        output_dir = os.path.expanduser("~/Desktop")
        base       = filename_base or f"leads_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}"
        csv_path   = os.path.join(output_dir, f"{base}.csv")
        excel_path = os.path.join(output_dir, f"{base}.xlsx")

        cleaned = [{k: clean_text(str(v)) for k, v in lead.items()} for lead in self.leads]
        df      = pd.DataFrame(cleaned)

        before = len(df)
        df     = df.drop_duplicates(subset=["business_name", "phone"])
        df     = df.reset_index(drop=True)

        order = ["source","business_name","email","phone","address","website","category","rating"]
        df    = df[[c for c in order if c in df.columns]]

        df.to_csv(csv_path, index=False, encoding="utf-8-sig")

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        col_labels = {
            "source":"Source","business_name":"Business Name",
            "email":"Email","phone":"Phone","address":"Address",
            "website":"Website","category":"Category","rating":"Rating",
        }

        header_fill = PatternFill("solid", fgColor="1C2B3A")
        odd_fill    = PatternFill("solid", fgColor="F7F7F5")
        even_fill   = PatternFill("solid", fgColor="FFFFFF")
        header_font = Font(bold=True, color="C9A84C", size=10, name="Calibri")
        body_font   = Font(size=9, name="Calibri", color="0D0D0D")
        email_font  = Font(size=9, name="Calibri", color="1A7A4A", underline="single")
        c_align     = Alignment(horizontal="center", vertical="center", wrap_text=False)
        l_align     = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        border      = Border(
            bottom=Side(style="thin", color="DDDDDD"),
            right =Side(style="thin", color="DDDDDD"),
        )

        columns = list(df.columns)
        ws.row_dimensions[1].height = 22

        for ci, col in enumerate(columns, 1):
            cell           = ws.cell(row=1, column=ci)
            cell.value     = col_labels.get(col, col.replace("_"," ").title())
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = c_align
            cell.border    = border

        for ri, row in df.iterrows():
            er   = ri + 2
            fill = odd_fill if ri % 2 == 0 else even_fill
            ws.row_dimensions[er].height = 18
            for ci, col in enumerate(columns, 1):
                cell           = ws.cell(row=er, column=ci)
                cell.value     = row[col]
                cell.fill      = fill
                cell.border    = border
                if col == "email" and row[col] != "N/A":
                    cell.font      = email_font
                    cell.alignment = l_align
                elif col in ("rating", "source"):
                    cell.font      = body_font
                    cell.alignment = c_align
                else:
                    cell.font      = body_font
                    cell.alignment = l_align

        for ci, col in enumerate(columns, 1):
            w = min(50, max(12, len(col_labels.get(col, col)) + 4,
                    df[col].astype(str).str.len().max() + 2 if len(df) else 0))
            if col in ("address","website"):
                w = min(45, w)
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(excel_path)

        has_email = len(df[df["email"] != "N/A"]) if "email" in df.columns else 0
        print(f"\n{'='*52}")
        print(f"  ✅  {len(df)} leads  ({before - len(df)} duplicates removed)")
        print(f"  📧  {has_email} leads with emails")
        print(f"  📊  Excel  → {excel_path}")
        print(f"  📄  CSV    → {csv_path}")
        print(f"{'='*52}")
        return excel_path, csv_path


# ─────────────────────────────────────────────────────────────
#  ██████╗ ██████╗ ███╗   ██╗████████╗██████╗  ██████╗ ██╗
# ██╔════╝██╔═══██╗████╗  ██║╚══██╔══╝██╔══██╗██╔═══██╗██║
# ██║     ██║   ██║██╔██╗ ██║   ██║   ██████╔╝██║   ██║██║
# ██║     ██║   ██║██║╚██╗██║   ██║   ██╔══██╗██║   ██║██║
# ╚██████╗╚██████╔╝██║ ╚████║   ██║   ██║  ██║╚██████╔╝███████╗
#  ╚═════╝ ╚═════╝ ╚═╝  ╚═══╝   ╚═╝   ╚═╝  ╚═╝ ╚═════╝ ╚══════╝
#
#   CHANGE ONLY THIS SECTION FOR EACH CLIENT
# ─────────────────────────────────────────────────────────────
async def main():

    # ── 1. WHAT TO SEARCH ────────────────────────────────────
    SEARCH_QUERY    = "digital marketing agency"

    # ── 2. WHERE ─────────────────────────────────────────────
    SEARCH_LOCATION = "Mumbai"

    # ── 3. HOW MANY LEADS ────────────────────────────────────
    MAX_RESULTS     = 20

    # ── 4. WHICH SOURCES TO USE ──────────────────────────────
    #   True  = scrape this site
    #   False = skip this site
    USE_GOOGLE_MAPS = True
    USE_JUSTDIAL    = True    # best for Indian businesses
    USE_INDIAMART   = False   # best for B2B / manufacturers only

    # ── 5. EMAIL ENRICHMENT ──────────────────────────────────
    #   Get free key at hunter.io (25 free/month)
    #   Leave blank "" to skip email finding
    HUNTER_API_KEY  = ""

    # ── 6. SHOW BROWSER WINDOW? ──────────────────────────────
    #   True  = browser opens visibly (good for debugging)
    #   False = runs silently in background
    SHOW_BROWSER    = True

    # ─────────────────────────────────────────────────────────
    # DO NOT CHANGE ANYTHING BELOW THIS LINE
    # ─────────────────────────────────────────────────────────
    scraper = LeadScraper(headless=not SHOW_BROWSER)

    if USE_GOOGLE_MAPS:
        await scraper.scrape_google_maps(SEARCH_QUERY, SEARCH_LOCATION, MAX_RESULTS)

    if USE_JUSTDIAL:
        await scraper.scrape_justdial(SEARCH_QUERY, SEARCH_LOCATION, MAX_RESULTS)

    if USE_INDIAMART:
        await scraper.scrape_indiamart(SEARCH_QUERY, SEARCH_LOCATION, MAX_RESULTS)

    if HUNTER_API_KEY and scraper.leads:
        await scraper.enrich_leads_with_emails(HUNTER_API_KEY)

    filename = f"{SEARCH_QUERY.replace(' ','_')}_{SEARCH_LOCATION}"
    scraper.export(filename)


if __name__ == "__main__":
    asyncio.run(main())