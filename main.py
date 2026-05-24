import asyncio
import os
import pandas as pd
from scrapper   import LeadScraper, clean_text
from lead_scorer import LeadScoringModel
from email_sender import ColdEmailCampaign
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
#   CONFIGURE EVERYTHING HERE — this is all you ever touch
# ═══════════════════════════════════════════════════════════════

# ── 1. WHAT TO SCRAPE ─────────────────────────────────────────
SEARCH_QUERY    = "digital marketing agency"
SEARCH_LOCATION = "Mumbai"
MAX_RESULTS     = 30

# ── 2. WHICH SOURCES ──────────────────────────────────────────
USE_GOOGLE_MAPS = True
USE_JUSTDIAL    = True
USE_INDIAMART   = False

# ── 3. EMAIL ENRICHMENT ───────────────────────────────────────
HUNTER_API_KEY  = ""           # hunter.io — free 25/month

# ── 4. YOUR DETAILS (for cold emails) ────────────────────────
YOUR_NAME       = "Sidharth Malik"
YOUR_COMPANY    = "Your Agency Name"
YOUR_SERVICE    = "help businesses grow with AI-powered marketing automation"
YOUR_WEBSITE    = "https://yourwebsite.com"

# ── 5. API KEYS ───────────────────────────────────────────────
ANTHROPIC_KEY   = ""           # anthropic.com — for email writing
GMAIL_ADDRESS   = ""           # your@gmail.com
GMAIL_APP_PASS  = ""           # myaccount.google.com → App Passwords

# ── 6. CAMPAIGN SETTINGS ─────────────────────────────────────
MIN_GRADE       = "B"          # Only email A and B grade leads
DRY_RUN         = True         # True = preview only, False = actually send
DAILY_LIMIT     = 20           # Max emails per run
SHOW_BROWSER    = True         # True = see the browser working

# ═══════════════════════════════════════════════════════════════


def export_scored_leads(df: pd.DataFrame, filename_base: str):
    """Export scored leads to Excel with color-coded grades."""
    output_dir = os.path.expanduser("~/Desktop")
    excel_path = os.path.join(output_dir, f"{filename_base}_SCORED.xlsx")
    csv_path   = os.path.join(output_dir, f"{filename_base}_SCORED.csv")

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    wb = Workbook()
    ws = wb.active
    ws.title = "Scored Leads"

    GRADE_COLORS = {
        "A": "1A7A4A",  # Green
        "B": "2980B9",  # Blue
        "C": "E67E22",  # Orange
        "D": "C0392B",  # Red
    }

    col_labels = {
        "ml_grade": "Grade", "ml_score": "Score",
        "business_name": "Business", "email": "Email",
        "phone": "Phone", "address": "Address",
        "website": "Website", "category": "Category",
        "rating": "Rating", "source": "Source",
        "ml_recommend": "Recommend?",
    }

    # Put important columns first
    priority = ["ml_grade","ml_score","ml_recommend","business_name",
                "email","phone","address","website","category","rating","source"]
    columns  = [c for c in priority if c in df.columns]

    header_fill = PatternFill("solid", fgColor="1C2B3A")
    header_font = Font(bold=True, color="C9A84C", size=10, name="Calibri")
    body_font   = Font(size=9, name="Calibri")
    c_align     = Alignment(horizontal="center", vertical="center")
    l_align     = Alignment(horizontal="left",   vertical="center")
    border      = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right =Side(style="thin", color="DDDDDD"),
    )

    ws.row_dimensions[1].height = 22
    for ci, col in enumerate(columns, 1):
        cell           = ws.cell(row=1, column=ci)
        cell.value     = col_labels.get(col, col.replace("_"," ").title())
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = c_align
        cell.border    = border

    for ri, row in df.iterrows():
        er   = ri + 2
        grade = str(row.get("ml_grade", "D"))
        row_bg = {
            "A": "EAF7EE", "B": "EBF5FB",
            "C": "FEF9E7", "D": "FDEDEC"
        }.get(grade, "FFFFFF")
        fill = PatternFill("solid", fgColor=row_bg)
        ws.row_dimensions[er].height = 18

        for ci, col in enumerate(columns, 1):
            cell       = ws.cell(row=er, column=ci)
            val        = row[col]
            cell.fill  = fill
            cell.border = border

            if col == "ml_grade":
                cell.value     = val
                cell.font      = Font(bold=True, size=11, name="Calibri",
                                      color=GRADE_COLORS.get(grade, "000000"))
                cell.alignment = c_align

            elif col == "ml_score":
                cell.value     = float(val)
                cell.font      = Font(bold=True, size=9, name="Calibri")
                cell.alignment = c_align

            elif col == "ml_recommend":
                cell.value     = "✅ YES" if val else "❌ NO"
                cell.font      = body_font
                cell.alignment = c_align

            elif col == "email" and str(val) != "N/A":
                cell.value     = val
                cell.font      = Font(size=9, name="Calibri",
                                      color="1A7A4A", underline="single")
                cell.alignment = l_align

            else:
                cell.value     = val
                cell.font      = body_font
                cell.alignment = c_align if col in ("source","rating") else l_align

    for ci, col in enumerate(columns, 1):
        w = min(50, max(12,
            len(col_labels.get(col, col)) + 4,
            df[col].astype(str).str.len().max() + 2 if len(df) else 0
        ))
        if col in ("address", "website"):
            w = min(40, w)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)

    print(f"\n[Export] Scored leads saved:")
    print(f"  📊 Excel → {excel_path}")
    print(f"  📄 CSV   → {csv_path}")
    return excel_path


async def main():
    base_name = f"{SEARCH_QUERY.replace(' ','_')}_{SEARCH_LOCATION}"

    # ── STEP 1: SCRAPE ────────────────────────────────────────
    print("\n" + "█"*55)
    print("  STEP 1 — SCRAPING LEADS")
    print("█"*55)

    scraper = LeadScraper(headless=not SHOW_BROWSER)

    if USE_GOOGLE_MAPS:
        await scraper.scrape_google_maps(SEARCH_QUERY, SEARCH_LOCATION, MAX_RESULTS)
    if USE_JUSTDIAL:
        await scraper.scrape_justdial(SEARCH_QUERY, SEARCH_LOCATION, MAX_RESULTS)
    if USE_INDIAMART:
        await scraper.scrape_indiamart(SEARCH_QUERY, SEARCH_LOCATION, MAX_RESULTS)

    if not scraper.leads:
        print("\n✗ No leads scraped. Check your settings and try again.")
        return

    # ── STEP 2: ENRICH WITH EMAILS ───────────────────────────
    if HUNTER_API_KEY:
        print("\n" + "█"*55)
        print("  STEP 2 — EMAIL ENRICHMENT")
        print("█"*55)
        await scraper.enrich_leads_with_emails(HUNTER_API_KEY)

    # ── STEP 3: ML SCORING ───────────────────────────────────
    print("\n" + "█"*55)
    print("  STEP 3 — ML LEAD SCORING")
    print("█"*55)

    df    = pd.DataFrame(scraper.leads)
    model = LeadScoringModel()
    model.train(df)                  # Train fresh on scraped data
    scored_df = model.score(df)      # Score all leads

    # Export scored Excel file
    export_scored_leads(scored_df, base_name)

    # ── STEP 4: COLD EMAIL CAMPAIGN ──────────────────────────
    if ANTHROPIC_KEY and GMAIL_ADDRESS and GMAIL_APP_PASS:
        print("\n" + "█"*55)
        print("  STEP 4 — COLD EMAIL CAMPAIGN")
        print("█"*55)

        campaign = ColdEmailCampaign(
            anthropic_api_key  = ANTHROPIC_KEY,
            gmail_address      = GMAIL_ADDRESS,
            gmail_app_password = GMAIL_APP_PASS,
            your_name          = YOUR_NAME,
            your_company       = YOUR_COMPANY,
            your_service       = YOUR_SERVICE,
            your_website       = YOUR_WEBSITE,
        )

        campaign.run(
            scored_df     = scored_df,
            min_grade     = MIN_GRADE,
            dry_run       = DRY_RUN,
            delay_seconds = (45, 90),
            daily_limit   = DAILY_LIMIT,
        )
    else:
        print("\n[Step 4] Skipped — add ANTHROPIC_KEY + Gmail credentials to send emails")
        print("         Scored leads exported to Desktop for manual review")

    print("\n✅ Full pipeline complete.")


if __name__ == "__main__":
    asyncio.run(main())